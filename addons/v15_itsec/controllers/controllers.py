from odoo import http
from odoo.http import request
import openpyxl
import io

class AuditTemplateController(http.Controller):

    @http.route('/download/template_audit_log', type='http', auth='user')
    def download_audit_template(self, **kwargs):
        # Header kolom
        headers = ['comp_name', 'standard', 'audit_date', 'mandays', 'role', 'audit_type', 'ea_code','audit_behalf',]

        # Buat workbook dan worksheet
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Audit Template"
        sheet.append(headers)  # Tambah header ke baris pertama

        # Simpan file ke memory (binary)
        file_stream = io.BytesIO()
        workbook.save(file_stream)
        file_stream.seek(0)  # Reset pointer ke awal file

        # Buat response untuk download file
        return request.make_response(
            file_stream.read(),
            headers=[
                ('Content-Type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'),
                ('Content-Disposition', 'attachment; filename="template_audit_log.xlsx"')
            ]
        )

class AuditorSkillTemplateController(http.Controller):

    @http.route('/download/template_audit_skill', type='http', auth='user')
    def download_audit_template(self, **kwargs):
        # Header kolom
        headers = ['iso_standard_ids', 'ea_code', 'auditor', 'lead_auditor', 'pk', 'keterangan']

        # Buat workbook dan worksheet
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Auditor Skill Template"
        sheet.append(headers)  # Tambah header ke baris pertama

        # Simpan file ke memory (binary)
        file_stream = io.BytesIO()
        workbook.save(file_stream)
        file_stream.seek(0)  # Reset pointer ke awal file

        # Buat response untuk download file
        return request.make_response(
            file_stream.read(),
            headers=[
                ('Content-Type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'),
                ('Content-Disposition', 'attachment; filename="template_audit_skill.xlsx"')
            ]
        )

class QMSExtraTemplateController(http.Controller):

    @http.route('/download/template_qms', type='http', auth='user')
    def download_template(self, type, **kwargs):
        # Buat header berbeda berdasarkan type
        if type == 'audit':
            headers = ['comp_name', 'standard', 'audit_date', 'mandays', 'role', 'audit_type', 'ea_code', 'audit_behalf']
            filename = 'template_audit.xlsx'
        elif type == 'skill':
            headers = ['iso_standard_ids', 'ea_code', 'auditor', 'lead_auditor', 'pk', 'keterangan']
            filename = 'template_skill.xlsx'
        elif type == 'training' or type == 'experience' or type == 'education':
            headers = ['periode', 'name']
            filename = f'template_{type}.xlsx'
        else:
            headers = []
            filename = 'template.xlsx'

        # Generate workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(headers)

        file_stream = io.BytesIO()
        wb.save(file_stream)
        file_stream.seek(0)

        return request.make_response(
            file_stream.read(),
            headers=[
                ('Content-Type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'),
                ('Content-Disposition', f'attachment; filename="{filename}"')
            ]
        )