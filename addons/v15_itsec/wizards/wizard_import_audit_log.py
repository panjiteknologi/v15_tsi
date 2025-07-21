from odoo import models, fields, api
import base64
import io
import openpyxl
from odoo.exceptions import UserError

class WizardAuditGabungan(models.TransientModel):
    _name = 'tsi.wizard.personal.information'
    _description = 'Import Wizard Personal Information'

    file = fields.Binary(string='File (.xlsx)')
    filename = fields.Char("Filename")
    employee_id = fields.Many2one('hr.employee', string="Employee")
    data_type = fields.Selection([
        ('audit', 'Audit'),
        ('training', 'Training'),
        ('experience', 'Experience'),
        ('education', 'Education'),
        ('skill', 'Skill')
    ], string="Data Type", required=True)

    def action_download_template(self):
        from openpyxl import Workbook
        import base64

        wb = Workbook()
        ws = wb.active

        if self.data_type == 'audit':
            ws.append(['Client Name', 'Standard', 'Audit Date', 'Mandays', 'Role', 'Audit Type', 'EA Code', 'Audit on Behalf'])
        elif self.data_type == 'skill':
            ws.append(['ISO Standard', 'EA Code', 'Auditor', 'Lead Auditor', 'PK', 'Keterangan'])
        elif self.data_type == 'education' or self.data_type == 'training' or self.data_type == 'experience':
            ws.append(['Periode', 'Name'])

        fp = io.BytesIO()
        wb.save(fp)
        fp.seek(0)
        file_data = base64.b64encode(fp.read())
        fp.close()

        self.write({
            'file': file_data,
            'filename': f"{self.data_type}_template.xlsx"
        })

        return {
            'type': 'ir.actions.act_url',
            'url': f'/download/template_qms?type={self.data_type}',
            'target': 'self',
        }

    def action_import_file(self):
        decoded_file = base64.b64decode(self.file)
        file_stream = io.BytesIO(decoded_file)
        workbook = openpyxl.load_workbook(file_stream)
        sheet = workbook.active

        for row in sheet.iter_rows(min_row=2, values_only=True):
            if self.data_type == 'audit':
                ea_code = self.env['tsi.employee.ea_code'].search([('name', '=', row[6])], limit=1)
                self.env['itsec.emp_audit'].create({
                    'emp_audit_id': self.employee_id.id,
                    'comp_name': row[0],
                    'standard': row[1],
                    'audit_date': row[2],
                    'mandays': row[3],
                    'role': row[4],
                    'audit_type': row[5],
                    'ea_code': [(6, 0, [ea_code.id])],
                    'audit_behalf': row[7],
                })
            elif self.data_type == 'skill':
                iso = self.env['tsi.standard.employee'].search([('name', '=', row[0])], limit=1)
                ea_code = self.env['tsi.employee.ea_code'].search([('name', '=', row[1])], limit=1)
                self.env['itsec.qms_extra'].create({
                    'qms_extra_id6': self.employee_id.id,
                    'iso_standard_ids': [(6, 0, [iso.id])],
                    'ea_code': [(6, 0, [ea_code.id])],
                    'auditor': row[2],
                    'lead_auditor': row[3],
                    'pk': row[4],
                    'keterangan': row[5],
                })
            else:  # education / training / experience
                mapping = {
                    'training': 'qms_extra_id3',
                    'experience': 'qms_extra_id4',
                    'education': 'qms_extra_id5'
                }
                field = mapping[self.data_type]
                self.env['itsec.qms_extra'].create({
                    field: self.employee_id.id,
                    'periode': row[0],
                    'name': row[1],
                })

        return {'type': 'ir.actions.act_window_close'}

class WizardAuditLog(models.TransientModel):
    _name = 'tsi.wizard.audit.log'
    _description = 'Import Wizard Audit Log'

    file = fields.Binary(string='File (.xlsx)')
    emp_audit_id = fields.Many2one('hr.employee', string="Employee")

    def action_import(self):
        # Decode the binary file
        decoded_file = base64.b64decode(self.file)
        
        # Open the file with openpyxl (binary mode)
        file_stream = io.BytesIO(decoded_file)
        workbook = openpyxl.load_workbook(file_stream)
        sheet = workbook.active
        
        # Process each row in the worksheet
        for row in sheet.iter_rows(min_row=2, values_only=True):  # Skip header row
            ea_code_name = row[6]

            # Cari EA Code (Many2many)
            ea_code = self.env['tsi.employee.ea_code'].search([('name', '=', ea_code_name)], limit=1)
            if not ea_code:
                raise UserError(f'EA Code \"{ea_code_name}\" tidak ditemukan di master data.')

            self.env['itsec.emp_audit'].create({
                'emp_audit_id': self.emp_audit_id.id,
                'comp_name': row[0],  # Second column
                'standard': row[1],  # Fourth column
                'audit_date': row[2],  # Assuming 'nomor' is the first column
                'mandays': row[3],  # Third column
                'role': row[4],  # Fifth column
                'audit_type': row[5],  # Sixth column
                'ea_code': [(6, 0, [ea_code.id])],
                'audit_behalf': row[7],  # Sixth column
            })
        return {'type': 'ir.actions.act_window_close'}

    def download_template(self):
        output = io.BytesIO()  # Use BytesIO for binary data
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(['nomor', 'comp_name', 'scope', 'standard', 'ea_code', 'role'])
        
        # Save to the output stream
        workbook.save(output)
        output.seek(0)
        
        # Encode the file for download
        file_data = base64.b64encode(output.read())

        return {
            'type': 'ir.actions.act_url',
            'url': '/web/content/?model=tsi.wizard.audit.log&id=%s&field=file&download=true&filename=audit_template.xlsx' % self.id,
            'target': 'self',
        }


class WizardAuditorSkill(models.TransientModel):
    _name = 'tsi.wizard.audit.skill'
    _description = 'Import Wizard Audit Skill'

    file = fields.Binary(string='File (.xlsx)')
    qms_extra_id6 = fields.Many2one('hr.employee', string="Employee")

    def action_import(self):
        decoded_file = base64.b64decode(self.file)
        file_stream = io.BytesIO(decoded_file)
        workbook = openpyxl.load_workbook(file_stream)
        sheet = workbook.active

        for row in sheet.iter_rows(min_row=2, values_only=True):  # Skip header
            iso_name = row[0]
            ea_code_name = row[1]

            # Cari ISO Standard (Many2many)
            iso_standard = self.env['tsi.standard.employee'].search([('name', '=', iso_name)], limit=1)
            if not iso_standard:
                raise UserError(f'Standard ISO \"{iso_name}\" tidak ditemukan di master data.')

            # Cari EA Code (Many2many)
            ea_code = self.env['tsi.employee.ea_code'].search([('name', '=', ea_code_name)], limit=1)
            if not ea_code:
                raise UserError(f'EA Code \"{ea_code_name}\" tidak ditemukan di master data.')

            self.env['itsec.qms_extra'].create({
                'qms_extra_id6': self.qms_extra_id6.id,
                'iso_standard_ids': [(6, 0, [iso_standard.id])],
                'ea_code': [(6, 0, [ea_code.id])],
                'auditor': row[2],
                'lead_auditor': row[3],
                'pk': row[4],
                'keterangan': row[5],
            })

        return {'type': 'ir.actions.act_window_close'}

    def download_template(self):
        output = io.BytesIO()  # Use BytesIO for binary data
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(['iso_standard_ids', 'ea_code', 'auditor', 'lead_auditor', 'pk', 'keterangan'])
        
        # Save to the output stream
        workbook.save(output)
        output.seek(0)
        
        # Encode the file for download
        file_data = base64.b64encode(output.read())

        return {
            'type': 'ir.actions.act_url',
            'url': '/web/content/?model=tsi.wizard.audit.log&id=%s&field=file&download=true&filename=auditor_skill_template.xlsx' % self.id,
            'target': 'self',
        }